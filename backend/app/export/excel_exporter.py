from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.export.base_exporter import BaseExporter
from app.export.contract_helpers import (
    get_export_data,
    get_export_errors,
    get_export_execution_summary,
    normalize_export_contract,
)


class ExcelExporter(BaseExporter):
    async def export(
        self,
        processed_data: dict[str, Any],
        *,
        analysis_data: dict[str, Any] | None = None,
        export_id: str | None = None,
        source_url: str = "",
        generated_at: str | None = None,
        title: str = "",
    ) -> str:
        resolved_export_id = self.resolve_export_id(export_id)
        contract = normalize_export_contract(processed_data, analysis_data=analysis_data, source_url=source_url)

        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = "Data"
        self._write_data_sheet(data_sheet, records=get_export_data(contract), title=title)

        execution_sheet = workbook.create_sheet(title="Execution")
        self._write_key_value_sheet(execution_sheet, "Execution Summary", get_export_execution_summary(contract))

        errors_sheet = workbook.create_sheet(title="Errors")
        self._write_errors_sheet(errors_sheet, get_export_errors(contract))

        buffer = BytesIO()
        workbook.save(buffer)
        return self.write_export_file("xlsx", buffer.getvalue(), export_id=resolved_export_id)

    def _write_data_sheet(
        self,
        worksheet: Any,
        *,
        records: list[dict[str, Any]],
        title: str,
    ) -> None:
        worksheet["A1"] = title or "Extracted Data"
        worksheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        worksheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")

        if not records:
            worksheet["A3"] = "No structured data available."
            worksheet.column_dimensions["A"].width = 28
            return

        headers = sorted({key for record in records for key in record.keys()}) or ["value"]
        for column_index, column_name in enumerate(headers, start=1):
            cell = worksheet.cell(row=3, column=column_index, value=column_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
            cell.alignment = Alignment(horizontal="center")

        for row_index, row_data in enumerate(records, start=4):
            for column_index, header in enumerate(headers, start=1):
                worksheet.cell(row=row_index, column=column_index, value=str(row_data.get(header, "")))

        for idx in range(1, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(idx)].width = 24

    def _write_key_value_sheet(self, worksheet: Any, title: str, values: dict[str, Any]) -> None:
        worksheet["A1"] = title
        worksheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        worksheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")

        current_row = 3
        for label, value in values.items():
            worksheet.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            worksheet.cell(row=current_row, column=2, value=value)
            worksheet.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            current_row += 1

        worksheet.column_dimensions["A"].width = 26
        worksheet.column_dimensions["B"].width = 60

    def _write_errors_sheet(self, worksheet: Any, errors: list[str]) -> None:
        worksheet["A1"] = "Errors"
        worksheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        worksheet["A1"].fill = PatternFill("solid", fgColor="7F1D1D")
        worksheet["A3"] = "Message"
        worksheet["A3"].font = Font(bold=True, color="FFFFFF")
        worksheet["A3"].fill = PatternFill("solid", fgColor="C0504D")

        for row_index, error in enumerate(errors, start=4):
            worksheet.cell(row=row_index, column=1, value=error)

        worksheet.column_dimensions["A"].width = 120
